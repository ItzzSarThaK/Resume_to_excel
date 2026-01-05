import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import os
import sys
from datetime import datetime

# Configure Tesseract path for Windows (if needed)
if sys.platform == 'win32':
    # Common Tesseract installation paths on Windows
    possible_paths = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        r'C:\Users\{}\AppData\Local\Tesseract-OCR\tesseract.exe'.format(os.getenv('USERNAME')),
    ]
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break


class ResumeScanner:
    def __init__(self, root):
        self.root = root
        self.root.title("Resume Scanner - PDF to Excel Converter")
        self.root.geometry("900x850")
        self.root.configure(bg="#f5f7fa")
        self.root.resizable(True, True)
        self.root.minsize(700, 600)  # Set minimum window size for responsiveness
        
        # Center the window
        self.center_window()
        
        self.resume_paths = []  # Changed to list for multiple files
        self.output_path = None
        
        # Color scheme
        self.colors = {
            'primary': '#667eea',
            'primary_dark': '#5568d3',
            'secondary': '#764ba2',
            'success': '#10b981',
            'success_dark': '#059669',
            'danger': '#ef4444',
            'warning': '#f59e0b',
            'bg_main': '#f5f7fa',
            'bg_card': '#ffffff',
            'text_primary': '#1f2937',
            'text_secondary': '#6b7280',
            'border': '#e5e7eb'
        }
        
        self.setup_ui()
        
    def center_window(self):
        """Center the window on screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def create_button(self, parent, text, command, bg_color, hover_color, state=tk.NORMAL, icon=""):
        """Create a styled button with hover effect"""
        btn = tk.Button(
            parent,
            text=f"{icon} {text}" if icon else text,
            command=command,
            font=("Segoe UI", 11, "bold"),
            bg=bg_color,
            fg="white",
            padx=30,
            pady=12,
            cursor="hand2", 
            relief=tk.FLAT,
            bd=0,
            state=state,
            activebackground=hover_color,
            activeforeground="white"
        )
        
        def on_enter(e):
            if btn['state'] == tk.NORMAL:
                btn.config(bg=hover_color)
        
        def on_leave(e):
            if btn['state'] == tk.NORMAL:
                btn.config(bg=bg_color)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
        
    def setup_ui(self):
        # Main container
        main_container = tk.Frame(self.root, bg=self.colors['bg_main'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Header section with gradient effect
        header_frame = tk.Frame(
            main_container,
            bg=self.colors['primary'],
            height=180
        )
        header_frame.pack(fill=tk.X, padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Title section
        title_container = tk.Frame(header_frame, bg=self.colors['primary'])
        title_container.pack(expand=True)
        
        # Main title with icon
        title_label = tk.Label(
            title_container,
            text="📄 Resume Scanner",
            font=("Segoe UI", 32, "bold"),
            bg=self.colors['primary'],
            fg="white"
        )
        title_label.pack(pady=(20, 5))
        
        # Subtitle
        subtitle_label = tk.Label(
            title_container,
            text="Extract resume information and export to Excel seamlessly",
            font=("Segoe UI", 11),
            bg=self.colors['primary'],
            fg="#e0e7ff"
        )
        subtitle_label.pack(pady=(0, 20))
        
        # Content area with scrollable canvas
        # Create a frame to hold the canvas and scrollbar
        scroll_container = tk.Frame(main_container, bg=self.colors['bg_main'])
        scroll_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Create canvas for scrolling
        self.content_canvas = tk.Canvas(scroll_container, bg=self.colors['bg_main'], highlightthickness=0)
        content_scrollbar = ttk.Scrollbar(scroll_container, orient="vertical", command=self.content_canvas.yview)
        
        # Create scrollable frame inside canvas
        content_frame = tk.Frame(self.content_canvas, bg=self.colors['bg_main'])
        
        # Configure canvas scroll region
        def configure_scroll_region(event=None):
            self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))
        
        content_frame.bind("<Configure>", configure_scroll_region)
        
        # Create window in canvas for the scrollable frame
        canvas_window = self.content_canvas.create_window((0, 0), window=content_frame, anchor="nw")
        
        # Update canvas window width when canvas is resized
        def configure_canvas_window(event):
            canvas_width = event.width
            self.content_canvas.itemconfig(canvas_window, width=canvas_width)
            configure_scroll_region()
        
        self.content_canvas.bind('<Configure>', configure_canvas_window)
        self.content_canvas.configure(yscrollcommand=content_scrollbar.set)
        
        # Pack canvas and scrollbar
        self.content_canvas.pack(side="left", fill="both", expand=True)
        content_scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def on_mousewheel(event):
            self.content_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        self.content_canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Add padding to content frame
        content_wrapper = tk.Frame(content_frame, bg=self.colors['bg_main'])
        content_wrapper.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Card container for file selection
        card_frame = tk.Frame(
            content_wrapper,
            bg=self.colors['bg_card'],
            relief=tk.FLAT,
            bd=0
        )
        card_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 20))
        
        # File selection section
        file_section = tk.Frame(card_frame, bg=self.colors['bg_card'])
        file_section.pack(fill=tk.X, padx=30, pady=25)
        
        # File selection label
        file_label_title = tk.Label(
            file_section,
            text="📁 Select Resume File",
            font=("Segoe UI", 14, "bold"),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary'],
            anchor="w"
        )
        file_label_title.pack(fill=tk.X, pady=(0, 15))
        
        # Select button
        button_container = tk.Frame(file_section, bg=self.colors['bg_card'])
        button_container.pack(fill=tk.X, pady=(0, 15))
        
        # Button container with responsive layout
        btn_left_frame = tk.Frame(button_container, bg=self.colors['bg_card'])
        btn_left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        btn_right_frame = tk.Frame(button_container, bg=self.colors['bg_card'])
        btn_right_frame.pack(side=tk.RIGHT, padx=(10, 0))
        
        self.select_btn = self.create_button(
            btn_left_frame,
            "Browse & Select Resumes (PDF/DOCX) - Multiple Files",
            self.select_resume,
            self.colors['primary'],
            self.colors['primary_dark'],
            icon="🔍"
        )
        self.select_btn.pack(fill=tk.X, expand=True)
        
        # Add Clear button
        self.clear_btn = self.create_button(
            btn_right_frame,
            "Clear Selection",
            self.clear_selection,
            self.colors['danger'],
            "#dc2626",
            state=tk.DISABLED,
            
        )
        # Ensure white text color, even when disabled
        self.clear_btn.config(fg="white", disabledforeground="white")
        self.clear_btn.pack()
        
        # File info display area with scrollbar
        file_info_frame = tk.Frame(
            file_section,
            bg="#f9fafb",
            relief=tk.FLAT,
            bd=1,
            height=150
        )
        file_info_frame.pack(fill=tk.BOTH, expand=False, pady=(0, 15))
        file_info_frame.pack_propagate(False)
        
        # Create scrollable frame for multiple files
        self.file_display_canvas = tk.Canvas(file_info_frame, bg="#f9fafb", highlightthickness=0)
        scrollbar = ttk.Scrollbar(file_info_frame, orient="vertical", command=self.file_display_canvas.yview)
        scrollable_frame = tk.Frame(self.file_display_canvas, bg="#f9fafb")
        
        def update_scrollregion(event=None):
            self.file_display_canvas.configure(scrollregion=self.file_display_canvas.bbox("all"))
        
        scrollable_frame.bind("<Configure>", update_scrollregion)
        
        canvas_window = self.file_display_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        self.file_display_canvas.configure(yscrollcommand=scrollbar.set)
        
        # Bind canvas resize to update scrollable frame width
        def configure_scroll_region(event):
            canvas_width = event.width
            self.file_display_canvas.itemconfig(canvas_window, width=canvas_width)
            update_scrollregion()
        
        self.file_display_canvas.bind('<Configure>', configure_scroll_region)
        
        self.file_display_canvas.pack(side="left", fill="both", expand=True, padx=(15, 0), pady=15)
        scrollbar.pack(side="right", fill="y", pady=15)
        
        self.file_display_container = scrollable_frame
        
        # Initial file display
        self.file_icon_label = tk.Label(
            self.file_display_container,
            text="📄",
            font=("Segoe UI", 20),
            bg="#f9fafb",
            fg=self.colors['text_secondary']
        )
        self.file_icon_label.pack(side=tk.LEFT, padx=(0, 10))
        
        file_text_frame = tk.Frame(self.file_display_container, bg="#f9fafb")
        file_text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.file_status_label = tk.Label(
            file_text_frame,
            text="No files selected",
            font=("Segoe UI", 10),
            bg="#f9fafb",
            fg=self.colors['text_secondary'],
            anchor="w"
        )
        self.file_status_label.pack(fill=tk.X)
        
        self.file_label = tk.Label(
            file_text_frame,
            text="Please select one or more PDF or DOCX resume files",
            font=("Segoe UI", 9),
            bg="#f9fafb",
            fg="#9ca3af",
            anchor="w",
            wraplength=500
        )
        self.file_label.pack(fill=tk.X, pady=(2, 0))
        
        # Status section
        status_section = tk.Frame(card_frame, bg=self.colors['bg_card'])
        status_section.pack(fill=tk.X, padx=30, pady=(20, 25))
        
        # Status label with icon
        status_frame = tk.Frame(status_section, bg=self.colors['bg_card'])
        status_frame.pack(fill=tk.X)
        
        self.status_icon_label = tk.Label(
            status_frame,
            text="",
            font=("Segoe UI", 14),
            bg=self.colors['bg_card'],
            fg=self.colors['text_secondary']
        )
        self.status_icon_label.pack(side=tk.LEFT, padx=(0, 8))
        
        self.status_label = tk.Label(
            status_frame,
            text="Ready to scan",
            font=("Segoe UI", 10),
            bg=self.colors['bg_card'],
            fg=self.colors['text_secondary']
        )
        self.status_label.pack(side=tk.LEFT)
        
        # CONVERT BUTTON SECTION - Always visible at bottom of main container
        convert_section = tk.Frame(main_container, bg=self.colors['bg_main'])
        convert_section.pack(fill=tk.X, side=tk.BOTTOM, pady=(20, 30))
        
        # Convert button container
        convert_container = tk.Frame(convert_section, bg=self.colors['bg_main'])
        convert_container.pack(fill=tk.X, padx=40, pady=20)
        
        # Label above button
        convert_label = tk.Label(
            convert_container,
            text="🚀 Convert to Excel",
            font=("Segoe UI", 16, "bold"),
            bg=self.colors['bg_main'],
            fg=self.colors['primary'],
            anchor="center"
        )
        convert_label.pack(pady=(0, 15))
        
        # Create a large, prominent convert button (responsive)
        self.scan_btn = tk.Button(
            convert_container,
            text="📊 CONVERT ALL RESUMES TO EXCEL",
            command=self.scan_and_export,
            font=("Segoe UI", 14, "bold"),
            bg=self.colors['success'],
            fg="white",
            padx=40,
            pady=18,
            cursor="hand2",
            relief=tk.FLAT,
            bd=0,
            state=tk.DISABLED,
            activebackground=self.colors['success_dark'],
            activeforeground="white",
            disabledforeground="white",
            wraplength=600
        )
        
        # Add hover effect
        def on_enter_scan(e):
            if self.scan_btn['state'] == tk.NORMAL:
                self.scan_btn.config(bg=self.colors['success_dark'])
        
        def on_leave_scan(e):
            if self.scan_btn['state'] == tk.NORMAL:
                self.scan_btn.config(bg=self.colors['success'])
        
        self.scan_btn.bind("<Enter>", on_enter_scan)
        self.scan_btn.bind("<Leave>", on_leave_scan)
        self.scan_btn.pack(pady=(0, 10))
        
        # Help text below button
        help_label = tk.Label(
            convert_container,
            text="Select resume files above, then click this button to convert them to Excel",
            font=("Segoe UI", 10),
            bg=self.colors['bg_main'],
            fg=self.colors['text_secondary'],
            anchor="center"
        )
        help_label.pack(pady=(5, 0))
        
    def clear_selection(self):
        """Clear all selected files"""
        self.resume_paths = []
        self.update_file_display()
        self.scan_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)
        self.status_icon_label.config(text="")
        self.status_label.config(
            text="Ready to scan",
            fg=self.colors['text_secondary']
        )
    
    def update_file_display(self):
        """Update the file display area with selected files"""
        # Clear existing file display widgets (except the container)
        for widget in self.file_display_container.winfo_children():
            widget.destroy()
        
        # Update canvas scrollregion after widgets are added
        def update_canvas_scroll():
            self.file_display_canvas.update_idletasks()
            self.file_display_canvas.configure(scrollregion=self.file_display_canvas.bbox("all"))
        
        if not self.resume_paths:
            # Show default message
            file_icon_label = tk.Label(
                self.file_display_container,
                text="📄",
                font=("Segoe UI", 20),
                bg="#f9fafb",
                fg=self.colors['text_secondary']
            )
            file_icon_label.pack(side=tk.LEFT, padx=(0, 10))
            
            file_text_frame = tk.Frame(self.file_display_container, bg="#f9fafb")
            file_text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            self.file_status_label = tk.Label(
                file_text_frame,
                text="No files selected",
                font=("Segoe UI", 10),
                bg="#f9fafb",
                fg=self.colors['text_secondary'],
                anchor="w"
            )
            self.file_status_label.pack(fill=tk.X)
            
            self.file_label = tk.Label(
                file_text_frame,
                text="Please select one or more PDF or DOCX resume files",
                font=("Segoe UI", 9),
                bg="#f9fafb",
                fg="#9ca3af",
                anchor="w",
                wraplength=500
            )
            self.file_label.pack(fill=tk.X, pady=(2, 0))
            update_canvas_scroll()
        else:
            # Display all selected files
            total_size = 0
            for idx, file_path in enumerate(self.resume_paths):
                filename = os.path.basename(file_path)
                file_size = os.path.getsize(file_path) / 1024  # Size in KB
                total_size += file_size
                
                # Create file item frame
                file_item_frame = tk.Frame(self.file_display_container, bg="#f9fafb")
                file_item_frame.pack(fill=tk.X, pady=(0, 10) if idx < len(self.resume_paths) - 1 else 0)
                
                # File icon
                file_icon = tk.Label(
                    file_item_frame,
                    text="📄",
                    font=("Segoe UI", 16),
                    bg="#f9fafb",
                    fg=self.colors['success']
                )
                file_icon.pack(side=tk.LEFT, padx=(0, 10))
                
                # File info
                file_info_frame = tk.Frame(file_item_frame, bg="#f9fafb")
                file_info_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                
                file_name_label = tk.Label(
                    file_info_frame,
                    text=f"{idx + 1}. {filename}",
                    font=("Segoe UI", 10, "bold"),
                    bg="#f9fafb",
                    fg=self.colors['text_primary'],
                    anchor="w"
                )
                file_name_label.pack(fill=tk.X)
                
                file_size_label = tk.Label(
                    file_info_frame,
                    text=f"Size: {file_size:.1f} KB",
                    font=("Segoe UI", 9),
                    bg="#f9fafb",
                    fg="#9ca3af",
                    anchor="w"
                )
                file_size_label.pack(fill=tk.X, pady=(2, 0))
            
            # Update summary label
            summary_frame = tk.Frame(self.file_display_container, bg="#f9fafb")
            summary_frame.pack(fill=tk.X, pady=(10, 0))
            
            self.file_status_label = tk.Label(
                summary_frame,
                text=f"✅ {len(self.resume_paths)} file(s) selected • Total size: {total_size:.1f} KB",
                font=("Segoe UI", 10, "bold"),
                bg="#f9fafb",
                fg=self.colors['success'],
                anchor="w"
            )
            self.file_status_label.pack(fill=tk.X)
            
            self.file_label = tk.Label(
                summary_frame,
                text="Ready to scan and export all resumes to one Excel file",
                font=("Segoe UI", 9),
                bg="#f9fafb",
                fg=self.colors['text_secondary'],
                anchor="w"
            )
            self.file_label.pack(fill=tk.X, pady=(2, 0))
            
            # Update canvas scrollregion after all widgets are added
            self.root.after(10, update_canvas_scroll)
    
    def select_resume(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Resume Files (Multiple Selection Allowed)",
            filetypes=[
                ("PDF files", "*.pdf"),
                ("Word documents", "*.docx"),
                ("All files", "*.*")
            ]
        )
        
        if file_paths:
            # Add new files to the list (avoid duplicates)
            for file_path in file_paths:
                if file_path not in self.resume_paths:
                    self.resume_paths.append(file_path)
            
            # Update file display
            self.update_file_display()
            
            # Enable scan button and clear button
            self.scan_btn.config(state=tk.NORMAL)
            self.clear_btn.config(state=tk.NORMAL)
            
            # Update status
            self.status_icon_label.config(text="📋")
            self.status_label.config(
                text=f"{len(self.resume_paths)} file(s) selected",
                fg=self.colors['success']
            )
            
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF file"""
        text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            # If text extraction fails, try OCR
            print(f"Text extraction failed, trying OCR: {e}")
            try:
                images = convert_from_path(pdf_path)
                for image in images:
                    text += pytesseract.image_to_string(image) + "\n"
            except Exception as ocr_error:
                print(f"OCR failed: {ocr_error}")
                raise Exception("Could not extract text from PDF")
        return text
    
    def extract_text_from_docx(self, docx_path):
        """Extract text from DOCX file"""
        from docx import Document
        doc = Document(docx_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    
    def extract_information(self, text):
        """Extract structured information from resume text"""
        info = {
            "Name": "",
            "Email": "",
            "Phone": "",
            "Address": "",
            "Skills": [],
            "Projects": [],
            "Education": [],
            "Experience": [],
            "Summary": ""
        }
        
        lines = text.split('\n')
        text_lower = text.lower()
        
        # Extract Email
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        emails = re.findall(email_pattern, text)
        if emails:
            info["Email"] = emails[0]
        
        # Extract Phone
        phone_patterns = [
            r'(\+?\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\d{3}[-.\s]?\d{3}[-.\s]?\d{4}',
            r'\+?\d{10,15}'
        ]
        for pattern in phone_patterns:
            phones = re.findall(pattern, text)
            if phones:
                info["Phone"] = phones[0] if isinstance(phones[0], str) else ''.join(phones[0])
                break
        
        # Extract Name (usually first line or before email)
        for i, line in enumerate(lines[:10]):  # Check first 10 lines
            line_stripped = line.strip()
            if line_stripped and len(line_stripped) > 2:
                # Check if it looks like a name (not email, phone, or common words)
                if '@' not in line_stripped and not re.search(r'\d{3}', line_stripped):
                    if not any(word in line_stripped.lower() for word in ['resume', 'cv', 'curriculum', 'vitae']):
                        info["Name"] = line_stripped
                        break
        
        # Extract Skills (exclude projects)
        skill_keywords = [
            'skills', 'technical skills', 'competencies', 'expertise',
            'programming languages', 'technologies', 'tools'
        ]
        skills_section = False
        for i, line in enumerate(lines):
            line_lower = line.lower().strip()
            if any(keyword in line_lower for keyword in skill_keywords):
                skills_section = True
                continue
            if skills_section:
                if line.strip() and not line.strip().startswith(('education', 'experience', 'work', 'employment', 'project')):
                    # Extract skills (comma or bullet separated)
                    skills = re.split(r'[,•·\-\n]', line)
                    for skill in skills:
                        skill = skill.strip()
                        if skill and len(skill) > 1:
                            # Don't add if it looks like a project (contains project-related keywords or is too long)
                            if not any(proj_word in skill.lower() for proj_word in ['project', 'github', 'repository', 'repo', 'website', 'app', 'application']):
                                info["Skills"].append(skill)
                else:
                    skills_section = False
                    if len(info["Skills"]) > 0:
                        break
        
        # Extract Projects
        project_keywords = ['projects', 'project', 'portfolio', 'personal projects', 'side projects', 'key projects']
        projects_section = False
        project_lines = []
        for i, line in enumerate(lines):
            line_lower = line.lower().strip()
            if any(keyword in line_lower for keyword in project_keywords):
                projects_section = True
                continue
            if projects_section:
                if line.strip() and not line.strip().startswith(('education', 'experience', 'work', 'employment', 'skills', 'reference', 'certification', 'award')):
                    # Collect project lines
                    if line.strip() and len(line.strip()) > 3:  # Only meaningful lines
                        project_lines.append(line.strip())
                else:
                    projects_section = False
                    if project_lines:
                        # Split projects (projects are typically separated by numbers, bullets, or empty lines)
                        current_project = []
                        for proj_line in project_lines[:25]:  # Limit to 25 lines
                            # Check if it's a new project (starts with number, bullet, dash, or is very short after a longer line)
                            line_stripped = proj_line.strip()
                            if line_stripped:
                                # Check if this looks like a new project header
                                is_new_project = (line_stripped[0].isdigit() or 
                                                 line_stripped[0] in ['•', '-', '*', '.'] or
                                                 (line_stripped[0].isupper() and len(current_project) > 0 and len(line_stripped) < 60) or
                                                 (len(current_project) > 0 and len(line_stripped) < 30 and not line_stripped[0].islower()))
                                
                                if is_new_project and current_project:
                                    # Save previous project
                                    project_text = ' '.join(current_project).strip()
                                    if project_text:
                                        info["Projects"].append(project_text)
                                    current_project = [line_stripped]
                                else:
                                    current_project.append(line_stripped)
                        # Add the last project
                        if current_project:
                            project_text = ' '.join(current_project).strip()
                            if project_text:
                                info["Projects"].append(project_text)
                        break
        
        # Extract Education
        education_keywords = ['education', 'academic', 'qualification', 'degree', 'university', 'college']
        education_section = False
        for i, line in enumerate(lines):
            line_lower = line.lower().strip()
            if any(keyword in line_lower for keyword in education_keywords):
                education_section = True
                continue
            if education_section:
                if line.strip() and not line.strip().startswith(('experience', 'work', 'employment', 'skills')):
                    if any(word in line.lower() for word in ['bachelor', 'master', 'phd', 'degree', 'diploma', 'certificate']):
                        info["Education"].append(line.strip())
                else:
                    education_section = False
                    if len(info["Education"]) > 0:
                        break
        
        # Extract Experience
        experience_keywords = ['experience', 'employment', 'work history', 'career', 'professional']
        experience_section = False
        experience_lines = []
        for i, line in enumerate(lines):
            line_lower = line.lower().strip()
            if any(keyword in line_lower for keyword in experience_keywords):
                experience_section = True
                continue
            if experience_section:
                if line.strip() and not line.strip().startswith(('education', 'skills', 'reference')):
                    experience_lines.append(line.strip())
                else:
                    experience_section = False
                    if experience_lines:
                        info["Experience"] = experience_lines[:10]  # Limit to 10 lines
                        break
        
        # Extract Summary/Objective (usually at the beginning)
        summary_keywords = ['summary', 'objective', 'profile', 'about']
        for i, line in enumerate(lines[:15]):
            line_lower = line.lower().strip()
            if any(keyword in line_lower for keyword in summary_keywords):
                if i + 1 < len(lines):
                    info["Summary"] = lines[i + 1].strip() if lines[i + 1].strip() else ""
                    break
        
        return info
    
    def export_to_excel(self, info, output_path):
        """Export extracted information to Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resume Information"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Data style
        data_font = Font(size=11)
        
        # Write headers
        headers = ["Field", "Information"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row = 2
        
        # Basic Information
        basic_info = [
            ("Name", info["Name"]),
            ("Email", info["Email"]),
            ("Phone", info["Phone"]),
            ("Address", info["Address"])
        ]
        
        for field, value in basic_info:
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).font = data_font
            row += 1
        
        # Summary
        if info["Summary"]:
            ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
            ws.cell(row=row, column=2, value=info["Summary"]).font = data_font
            row += 1
        
        # Skills
        if info["Skills"]:
            ws.cell(row=row, column=1, value="Skills").font = Font(bold=True)
            skills_text = ", ".join(info["Skills"][:20])  # Limit to 20 skills
            ws.cell(row=row, column=2, value=skills_text).font = data_font
            row += 1
        
        # Projects
        if info["Projects"]:
            ws.cell(row=row, column=1, value="Projects").font = Font(bold=True)
            projects_text = "\n".join(info["Projects"][:10])  # Limit to 10 projects
            ws.cell(row=row, column=2, value=projects_text).font = data_font
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        # Education
        if info["Education"]:
            ws.cell(row=row, column=1, value="Education").font = Font(bold=True)
            education_text = "\n".join(info["Education"][:5])  # Limit to 5 entries
            ws.cell(row=row, column=2, value=education_text).font = data_font
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        # Experience
        if info["Experience"]:
            ws.cell(row=row, column=1, value="Experience").font = Font(bold=True)
            experience_text = "\n".join(info["Experience"][:10])  # Limit to 10 lines
            ws.cell(row=row, column=2, value=experience_text).font = data_font
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
        
        # Adjust row heights
        for row_num in range(1, row + 1):
            ws.row_dimensions[row_num].height = 25
        
        # Save workbook
        wb.save(output_path)
    
    def export_multiple_to_excel(self, all_resume_info, output_path):
        """Export multiple resumes' information to a single Excel file"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resume Information"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Data style
        data_font = Font(size=11)
        
        # Define column headers - each resume will be a row
        headers = [
            "File Name", "Name", "Email", "Phone", "Address", 
            "Summary", "Skills", "Projects", "Education", "Experience"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data for each resume
        for resume_idx, resume_data in enumerate(all_resume_info, start=2):
            info = resume_data['info']
            filename = resume_data['filename']
            
            # File Name
            ws.cell(row=resume_idx, column=1, value=filename).font = Font(bold=True, size=11)
            
            # Name
            ws.cell(row=resume_idx, column=2, value=info.get("Name", "")).font = data_font
            
            # Email
            ws.cell(row=resume_idx, column=3, value=info.get("Email", "")).font = data_font
            
            # Phone
            ws.cell(row=resume_idx, column=4, value=info.get("Phone", "")).font = data_font
            
            # Address
            ws.cell(row=resume_idx, column=5, value=info.get("Address", "")).font = data_font
            
            # Summary
            summary_text = info.get("Summary", "")
            ws.cell(row=resume_idx, column=6, value=summary_text).font = data_font
            ws.cell(row=resume_idx, column=6).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Skills
            skills_text = ", ".join(info.get("Skills", [])[:30])  # Limit to 30 skills
            ws.cell(row=resume_idx, column=7, value=skills_text).font = data_font
            ws.cell(row=resume_idx, column=7).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Projects
            projects_text = "\n\n".join(info.get("Projects", [])[:10])  # Limit to 10 projects, separated by double newline
            ws.cell(row=resume_idx, column=8, value=projects_text).font = data_font
            ws.cell(row=resume_idx, column=8).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Education
            education_text = "\n".join(info.get("Education", [])[:10])  # Limit to 10 entries
            ws.cell(row=resume_idx, column=9, value=education_text).font = data_font
            ws.cell(row=resume_idx, column=9).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Experience
            experience_text = "\n".join(info.get("Experience", [])[:15])  # Limit to 15 lines
            ws.cell(row=resume_idx, column=10, value=experience_text).font = data_font
            ws.cell(row=resume_idx, column=10).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Set row height for better visibility
            ws.row_dimensions[resume_idx].height = 80
        
        # Adjust column widths (updated to include Projects column)
        column_widths = [25, 20, 30, 15, 30, 40, 50, 50, 40, 60]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Freeze the header row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
    
    def scan_and_export(self):
        if not self.resume_paths:
            messagebox.showerror("Error", "Please select at least one resume file first!")
            return
        
        try:
            # Disable buttons during processing
            self.select_btn.config(state=tk.DISABLED)
            self.scan_btn.config(state=tk.DISABLED)
            self.clear_btn.config(state=tk.DISABLED)
            
            # Update status
            self.status_icon_label.config(text="⏳")
            self.status_label.config(
                text=f"Processing {len(self.resume_paths)} resume(s)...",
                fg=self.colors['primary']
            )
            self.root.update()
            
            all_resume_info = []
            
            # Process each resume
            for idx, resume_path in enumerate(self.resume_paths, 1):
                filename = os.path.basename(resume_path)
                
                # Update status for current file
                self.status_icon_label.config(text="⏳")
                self.status_label.config(
                    text=f"Processing {idx}/{len(self.resume_paths)}: {filename}...",
                    fg=self.colors['primary']
                )
                self.root.update()
                
                try:
                    # Extract text based on file type
                    if resume_path.lower().endswith('.pdf'):
                        text = self.extract_text_from_pdf(resume_path)
                    elif resume_path.lower().endswith('.docx'):
                        text = self.extract_text_from_docx(resume_path)
                    else:
                        # Skip unsupported files
                        continue
                    
                    if not text.strip():
                        # Skip files with no extractable text
                        continue
                    
                    # Extract information
                    info = self.extract_information(text)
                    
                    # Store resume info with filename
                    all_resume_info.append({
                        'filename': filename,
                        'info': info
                    })
                    
                except Exception as file_error:
                    # Continue processing other files even if one fails
                    print(f"Error processing {filename}: {file_error}")
                    continue
            
            if not all_resume_info:
                messagebox.showerror("Error", "Could not extract information from any of the selected files!")
                self.select_btn.config(state=tk.NORMAL)
                self.scan_btn.config(state=tk.NORMAL)
                self.clear_btn.config(state=tk.NORMAL)
                return
            
            self.status_icon_label.config(text="📊")
            self.status_label.config(
                text=f"Creating Excel file with {len(all_resume_info)} resume(s)...",
                fg=self.colors['primary']
            )
            self.root.update()
            
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # Use directory of first file, or current directory
            base_dir = os.path.dirname(self.resume_paths[0]) if self.resume_paths else os.getcwd()
            output_path = os.path.join(
                base_dir,
                f"All_Resumes_Extracted_{timestamp}.xlsx"
            )
            
            # Export all resumes to Excel
            self.export_multiple_to_excel(all_resume_info, output_path)
            
            # Success
            self.status_icon_label.config(text="✅")
            self.status_label.config(
                text=f"Success! {len(all_resume_info)} resume(s) exported to Excel",
                fg=self.colors['success']
            )
            
            # Re-enable buttons
            self.select_btn.config(state=tk.NORMAL)
            self.scan_btn.config(state=tk.NORMAL)
            self.clear_btn.config(state=tk.NORMAL)
            
            # Show success message with better formatting
            messagebox.showinfo(
                "✅ Success!",
                f"Resume information extracted successfully!\n\n"
                f"📊 {len(all_resume_info)} resume(s) processed\n"
                f"📁 File saved to:\n{output_path}\n\n"
                f"📋 All resumes are stored in one Excel file with each resume as a separate row."
            )
            
        except Exception as e:
            self.status_icon_label.config(text="❌")
            self.status_label.config(
                text=f"Error: {str(e)[:50]}...",
                fg=self.colors['danger']
            )
            
            # Re-enable buttons
            self.select_btn.config(state=tk.NORMAL)
            self.scan_btn.config(state=tk.NORMAL)
            self.clear_btn.config(state=tk.NORMAL)
            
            messagebox.showerror(
                "❌ Error",
                f"An error occurred while processing the resumes:\n\n{str(e)}"
            )


def main():
    root = tk.Tk()
    app = ResumeScanner(root)
    root.mainloop()


if __name__ == "__main__":
    main()

